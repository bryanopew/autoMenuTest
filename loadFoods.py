from openpyxl import load_workbook


def reloadFoods(foodList, remain, currentInfo):

    priceRemain = remain.get("priceRemain")
    calRemain = remain.get("calRemain")
    carbRemain = remain.get("carbRemain")
    protRemain = remain.get("protRemain")
    fatRemain = remain.get("fatRemain")
    categoryNumberRemain = remain.get("categoryNumberRemain")
    recommendedList = currentInfo.get("recommendedList")
    

    print("remain:", priceRemain, calRemain, carbRemain, protRemain, fatRemain)
    foodListMod = [[], [], [], [], [], []]

    for i, category in enumerate(foodList):
        # categoryNumberRemain이 다 0인데 목표섭취량 이하 식품은 있는 경우 처리를 위해
        # 식품을 먼저 받고 -> 다 0인경우 확인해서 categoryNumberRemain 수정 -> 카테고리0인경우 식품제거

        ####################### test1 #####################
        # if categoryNumberRemain[i] == 0:
        #     continue
        ###################################################
        for row in category:
            # 0: 카테고리 / 1: 순번 / 2: 식품명 / 3: 가격 / 4: 칼 / 5: 탄 / 6: 단 / 7: 지
            # print(row)
            if None in row:
                continue

            # 들어있는 상품 제외
            same = False
            for food in recommendedList:
                if food[0] == row[0] and food[1] == row[1]:
                    same = True
                    break
            if same:
                continue
            
            # if (
            #     row[3] <= priceRemain
            #     and row[4] <= calRemain + 50
            #     and row[5] <= carbRemain + 10
            #     and row[6] <= protRemain + 5
            #     and row[7] <= fatRemain + 3
            # ):
            #     foodListMod[i].append(row)

            # 이거때문에 많이느려지나...??!?! -> 영향크면 우리가 어떤 것 때문에 추천 힘든지만 파악하고
            # 실제로는 위 if 문으로 그냥 reload해주는게 좋을 것 (당장은 단백질이 문제인 듯) -> 10까지만 올려도 개빠름
            remainForCheck = [priceRemain, calRemain+50, carbRemain+10, protRemain+5, fatRemain+3]
            notSatisfiedTemp = [0, 0, 0, 0, 0]
            for j, value in enumerate(remainForCheck):
                if row[j+3] <= value:
                    notSatisfiedTemp[j] = 0
                else:
                    notSatisfiedTemp[j] = 1
            if 1 not in notSatisfiedTemp:
                foodListMod[i].append(row)
            else:
                for k, l in enumerate(notSatisfiedTemp):
                    # 여기서는 deepCopy 없이 -> 리턴안해줘도 값 바뀌어있을 것
                    currentInfo["notSatisfied"][k] += l


    # 목표섭취량이 높게 설정되어 있는 경우
    # 식품 reload할 때 식품은 있는데(목표영양은 남았는데) 
    # categoryNumberRemain이 부족한 경우 증가시키기
    foodExist = 0
    print("카테고리별 식품 수: ", end=" | ")
    for category in foodListMod:
        print(len(category), end=" | ")
        foodExist += len(category)
    print()

    zeroNum = 0
    for numberRemain in categoryNumberRemain:
        if numberRemain == 0:
            zeroNum += 1

    if zeroNum == len(foodListMod) and foodExist != 0:
        print("식품은 있는데 남은 numberRemain 하나도 없음")
        for i in range(len(foodListMod)):
            if len(foodListMod[i]) == 0:
                continue
            categoryNumberRemain[i] += 1
    else:
        for i, numberRemain in enumerate(categoryNumberRemain):
            if categoryNumberRemain[i] == 0:
                foodListMod[i] = []
    #############################################################
    
    return foodListMod


def loadFoods(remain, currentInfo):
    # 식품정보
    # [[name, cal, carb, prot, fat, price], [], ...] !!!!!

    # 기본 변수

    # 추천할 수 있는 식품유형 최대 개수
    # 도시락 / 닭가슴살 / 샐러드 / 간식 / 과자 / 음료
    categoryNumberMax = [1, 2, 1, 2, 2, 1]

    # 식품정보 가져오기 (일단 전체 정보)
    ws = load_workbook(
        "C:/Users/bryan/OneDrive/문서/두비/2022/005_식품DB/식품관리_v1.1_220713.xlsx"
    )
    dosiDb = ws["도시락"]
    chickenDb = ws["닭가슴살"]
    saladDb = ws["샐러드"]
    snackDb = ws["영양간식"]
    barDb = ws["과자"]
    drinkDb = ws["음료"]

    dosiNum = 79
    chickenNum = 70
    saladNum = 29
    snackNum = 20
    barNum = 23
    drinkNum = 17

    db = [dosiDb, chickenDb, saladDb, snackDb, barDb, drinkDb]
    foodsNumTotal = [dosiNum, chickenNum, saladNum, snackNum, barNum, drinkNum]

    # foodList -> 0: 도시락 / 1: 닭가슴살 / 2: 샐러드 / 3: 간식 / 4: 과자 / 5: 음료
    foodList = [[], [], [], [], [], []]

    # 엑셀데이터-> 1: 카테고리 / 3: 순번 / 4: 식품명 / 5: 가격 / 11~14: 칼탄단지
    food_temp = []

    for i, categoryNum in enumerate(foodsNumTotal):
        if remain.get("categoryNumberRemain")[i] <= 0:
            continue
        for j in range(2, categoryNum + 1):
            food_temp = []
            for k, cell in enumerate(db[i][j]):
                if k == 1 or k == 3 or k == 4 or k == 5 or (11 <= k and k <= 14) or k==29:
                    food_temp.append(cell.value)
            foodList[i].append(food_temp)
    foodListMod = reloadFoods(foodList, remain, currentInfo)
    return foodListMod
